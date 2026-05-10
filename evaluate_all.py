"""Comprehensive 4-Way Evaluation Script
Compares on the same environment and task set:
  1. C-LA-MAML        — dual adapters (lang_model/)
  2. Unified LA-MAML  — single adapter, full combined string (unified_model/)
  3. Hyper C-LA-MAML  — AbsoluteHyperNetwork (hyper_model/)
  4. Random Policy    — untrained baseline

Output: evaluation_results_all.xlsx  (one sheet per env, plus a per-mission sheet)
"""
import os, random, builtins, io
import warnings, logging
warnings.filterwarnings("ignore")
logging.getLogger("gymnasium").setLevel(logging.ERROR)

import torch
import numpy as np
from collections import OrderedDict
import argparse
from contextlib import contextmanager, redirect_stdout, redirect_stderr
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from torch.nn.utils.convert_parameters import parameters_to_vector

from environment import (ConstrainedGoToLocalEnv, ConstrainedPickupDistEnv,
                         ConstrainedGoToObjDoorEnv, ConstrainedGoToOpenEnv,
                         ConstrainedOpenDoorEnv, ConstrainedOpenDoorLocEnv,
                         ConstrainedOpenDoorsOrderEnv,
                         ConstrainedActionObjDoorEnv,
                         ConstrainedFindObjS5Env)
from sampler_lang import (BabyAIMissionTaskWrapper, SentenceMissionEncoder,
                          MissionParamAdapter, ConstraintParamAdapter,
                          AbsoluteHyperNetwork)
import sampler_lang
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy

# ── Helpers ────────────────────────────────────────────────────────────────────
@contextmanager
def silence():
    real_print = builtins.print
    buf = io.StringIO()
    def fp(*args, **kwargs):
        if args and isinstance(args[0], str) and "Sampling rejected" in args[0]:
            return
        real_print(*args, **kwargs)
    builtins.print = fp
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print

# ── CLI ────────────────────────────────────────────────────────────────────────
p = argparse.ArgumentParser()
p.add_argument("--env", dest="env_name",
               choices=["ConstrainedGoToLocal","ConstrainedPickupDist",
                        "ConstrainedGoToObjDoor","ConstrainedGoToOpen",
                        "ConstrainedOpenDoor","ConstrainedOpenDoorLoc",
                        "ConstrainedOpenDoorsOrder", "ConstrainedActionObjDoor",
                        "ConstrainedFindObjS5"],
               default="ConstrainedGoToLocal")
p.add_argument("--room-size",  type=int,   default=8)
p.add_argument("--num-dists",  type=int,   default=2)
p.add_argument("--max-steps",  type=int,   default=50)
p.add_argument("--delta-theta",type=float, default=0.3)
p.add_argument("--delta-constraint", type=float, default=0.1)
p.add_argument("--n-missions", type=int,   default=10,  help="tasks to evaluate per method")
p.add_argument("--n-episodes", type=int,   default=10,  help="episodes per task")
p.add_argument("--skip-clamaml",  action="store_true")
p.add_argument("--skip-unified",  action="store_true")
p.add_argument("--skip-hyper",    action="store_true")
p.add_argument("--skip-random",   action="store_true")
args = p.parse_args()

seed = 42
random.seed(seed); np.random.seed(seed); torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ── Eval mission lists (eval colors/objects differ from train) ─────────────────
OBJECTS     = ['box']
COLORS      = ['red','green','blue','purple','yellow','grey']
PREP_LOCS   = ['on','at','to']
LOC_NAMES   = ['right','front']
DOOR_COLORS = ['yellow','grey']

_LOCAL    = [f"go to the {c} {o}" for c in COLORS for o in OBJECTS]
_PICKUP   = [f"pick up the {c} {o}" for c in COLORS for o in OBJECTS]
_DOOR     = [f"go to the {c} door" for c in DOOR_COLORS]
_OPENDOOR = [f"open the {c} door" for c in DOOR_COLORS]
_DOORLOC  = [f"open the door {p} the {l}" for p in PREP_LOCS for l in LOC_NAMES]
_ORDERSEQ = (
    [f"open the {c} door" for c in DOOR_COLORS] +
    [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS] +
    [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)
_CT = [f"avoid {h}" for h in ['lava','grass','water']]
_DOUBLE_CT = [f"avoid {h1} and avoid {h2}" for i, h1 in enumerate(['lava','grass','water']) for h2 in ['lava','grass','water'][i+1:]]

_ACTIONOBJDOOR = (
    [f"pick up the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} door" for c in DOOR_COLORS] +
    [f"open the {c} door" for c in DOOR_COLORS]
)
_FINDOBJS5 = [f"pick up the {t}" for t in ["box"]]

MISSION_MAP = {
    "ConstrainedGoToLocal":     [f"{g} and {c}" for g in _LOCAL   for c in _DOUBLE_CT],
    "ConstrainedPickupDist":    [f"{g} and {c}" for g in _PICKUP  for c in _DOUBLE_CT],
    "ConstrainedGoToObjDoor":   [f"{g} and {c}" for g in _LOCAL+_DOOR for c in _DOUBLE_CT],
    "ConstrainedGoToOpen":      [f"{g} and {c}" for g in _LOCAL   for c in _DOUBLE_CT],
    "ConstrainedOpenDoor":      [f"{g} and {c}" for g in _OPENDOOR for c in _DOUBLE_CT],
    "ConstrainedOpenDoorLoc":   [f"{g} and {c}" for g in _OPENDOOR+_DOORLOC for c in _DOUBLE_CT],
    "ConstrainedOpenDoorsOrder":[f"{g} and {c}" for g in _ORDERSEQ for c in _DOUBLE_CT],
    "ConstrainedActionObjDoor": [f"{g} and {c}" for g in _ACTIONOBJDOOR for c in _DOUBLE_CT],
    "ConstrainedFindObjS5":     [f"{g} and {c}" for g in _FINDOBJS5 for c in _DOUBLE_CT],
}

GOALS_MAP = {
    "ConstrainedGoToLocal":      _LOCAL,
    "ConstrainedPickupDist":     _PICKUP,
    "ConstrainedGoToObjDoor":    _LOCAL + _DOOR,
    "ConstrainedGoToOpen":       _LOCAL,
    "ConstrainedOpenDoor":       _OPENDOOR,
    "ConstrainedOpenDoorLoc":    _OPENDOOR + _DOORLOC,
    "ConstrainedOpenDoorsOrder": _ORDERSEQ,
    "ConstrainedActionObjDoor":  _ACTIONOBJDOOR,
    "ConstrainedFindObjS5":      _FINDOBJS5,
}

# ── Build environment ──────────────────────────────────────────────────────────
def build_env(env_name, room_size, num_dists, max_steps, missions,
              goals=None, constraints=None):
    dispatch = {
        "ConstrainedGoToLocal":     lambda: ConstrainedGoToLocalEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedPickupDist":    lambda: ConstrainedPickupDistEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedGoToObjDoor":   lambda: ConstrainedGoToObjDoorEnv(max_steps=max_steps, num_distractors=num_dists),
        "ConstrainedGoToOpen":      lambda: ConstrainedGoToOpenEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedOpenDoor":      lambda: ConstrainedOpenDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorLoc":   lambda: ConstrainedOpenDoorLocEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorsOrder":lambda: ConstrainedOpenDoorsOrderEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedActionObjDoor": lambda: ConstrainedActionObjDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedFindObjS5":     lambda: ConstrainedFindObjS5Env(room_size=5, max_steps=max_steps),
    }
    base = dispatch[env_name]()
    return BabyAIMissionTaskWrapper(base, missions=missions, goals=goals, constraints=constraints)


# ── Setup: env + shared encoder ───────────────────────────────────────────────
env_name    = args.env_name
room_size   = "env" if env_name in ["GoToObjDoor"] else args.room_size
num_dists   = "env" if env_name in ["OpenDoor","OpenDoorLoc","OpenDoorsOrder"] else args.num_dists
max_steps   = args.max_steps
delta_theta = args.delta_theta
delta_c     = args.delta_constraint
N           = args.n_missions
EP          = args.n_episodes

all_missions   = MISSION_MAP[env_name]
goals_list     = GOALS_MAP.get(env_name)
constraints_list = _DOUBLE_CT if goals_list else None

env = build_env(env_name, room_size, num_dists, max_steps, all_missions,
                goals_list, constraints_list)

print(f"\n{'='*65}")
print(f"  4-Way Evaluation: {env_name}")
print(f"  Tasks: {N}  |  Episodes/task: {EP}  |  delta_theta: {delta_theta}")
print(f"{'='*65}\n")

dummy_obs, _ = env.reset()
input_size   = sampler_lang.preprocess_obs(dummy_obs).shape[0]
output_size  = env.action_space.n
hidden_sizes = (64, 64)
nl           = torch.nn.functional.tanh

mission_encoder = SentenceMissionEncoder(
    model_name="all-MiniLM-L6-v2", frozen=True, normalize=True, cache=True, device=device
)
mission_encoder.eval()
enc_dim = mission_encoder.output_dim

def _make_policy():
    pol = CategoricalMLPPolicy(input_size=input_size, output_size=output_size,
                               hidden_sizes=hidden_sizes, nonlinearity=nl).to(device)
    return pol

policy_param_shapes = [p.shape for p in _make_policy().parameters()]


# ── Load C-LA-MAML ─────────────────────────────────────────────────────────────
clamaml_ready = False
if not args.skip_clamaml:
    _ckpt_path = f"lang_model/lang_{env_name}_{delta_theta}.pth"
    if os.path.exists(_ckpt_path):
        ckpt_c = torch.load(_ckpt_path, map_location=device)
        policy_c = _make_policy(); policy_c.load_state_dict(ckpt_c["policy"]); policy_c.eval()
        adapter_c = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
        adapter_c.load_state_dict(ckpt_c["mission_adapter"]); adapter_c.eval()
        constr_adapter = None
        if "constraint_adapter" in ckpt_c:
            constr_adapter = ConstraintParamAdapter(enc_dim, policy_param_shapes).to(device)
            constr_adapter.load_state_dict(ckpt_c["constraint_adapter"]); constr_adapter.eval()
        clamaml_ready = True
        print(f"[✓] C-LA-MAML loaded from {_ckpt_path}")
    else:
        print(f"[✗] C-LA-MAML checkpoint not found: {_ckpt_path}  (skipping)")


# ── Load Unified LA-MAML ───────────────────────────────────────────────────────
unified_ready = False
if not args.skip_unified:
    _ckpt_path = f"unified_model/lang_{env_name}_{delta_theta}.pth"
    if os.path.exists(_ckpt_path):
        ckpt_u = torch.load(_ckpt_path, map_location=device)
        policy_u = _make_policy(); policy_u.load_state_dict(ckpt_u["policy"]); policy_u.eval()
        adapter_u = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
        adapter_u.load_state_dict(ckpt_u["mission_adapter"]); adapter_u.eval()
        unified_ready = True
        print(f"[✓] Unified LA-MAML loaded from {_ckpt_path}")
    else:
        print(f"[✗] Unified LA-MAML checkpoint not found: {_ckpt_path}  (skipping)")


# ── Load Hyper C-LA-MAML ───────────────────────────────────────────────────────
hyper_ready = False
if not args.skip_hyper:
    _ckpt_path = f"hyper_model/lang_{env_name}_{delta_theta}.pth"
    if os.path.exists(_ckpt_path):
        ckpt_h = torch.load(_ckpt_path, map_location=device)
        policy_h = _make_policy(); policy_h.load_state_dict(ckpt_h["policy"]); policy_h.eval()
        hypernetwork = None
        if ckpt_h.get("hypernetwork") is not None:
            hypernetwork = AbsoluteHyperNetwork(enc_dim, policy_param_shapes).to(device)
            hypernetwork.load_state_dict(ckpt_h["hypernetwork"]); hypernetwork.eval()
        hyper_ready = True
        print(f"[✓] Hyper C-LA-MAML loaded from {_ckpt_path}")
    else:
        print(f"[✗] Hyper C-LA-MAML checkpoint not found: {_ckpt_path}  (skipping)")

print()

# ── Adapted params helpers ─────────────────────────────────────────────────────
def _params_clamaml(mission):
    """θ' = θ + Δθ_goal + Δθ_constraint"""
    if isinstance(mission, tuple):
        goal_str, constr_str = mission
    else:
        parts = mission.split(" and avoid ", 1)
        goal_str, constr_str = (parts[0], "avoid " + parts[1]) if len(parts) == 2 else (mission, None)

    with torch.no_grad():
        g_emb   = mission_encoder(goal_str).to(device)
        deltas_g = adapter_c(g_emb)
        deltas_c  = [torch.zeros_like(d) for d in deltas_g]
        if constr_str and constr_adapter:
            c_emb    = mission_encoder(constr_str).to(device)
            deltas_c = constr_adapter(c_emb)

        names  = list(dict(policy_c.named_parameters()).keys())
        params = list(policy_c.parameters())
        return OrderedDict(
            (n, p + dg.squeeze(0) * delta_theta + dc.squeeze(0) * delta_c)
            for n, p, dg, dc in zip(names, params, deltas_g, deltas_c)
        )


def _params_unified(mission):
    """θ' = θ + Δθ_unified  (full combined string)"""
    combined = f"{mission[0]} and {mission[1]}" if isinstance(mission, tuple) else mission
    with torch.no_grad():
        emb    = mission_encoder(combined).to(device)
        deltas = adapter_u(emb)
        names  = list(dict(policy_u.named_parameters()).keys())
        params = list(policy_u.parameters())
        return OrderedDict(
            (n, p + d.squeeze(0) * delta_theta)
            for n, p, d in zip(names, params, deltas)
        )


def _params_hyper(mission):
    """θ' = HyperNetwork(θ, goal_emb, constr_emb)"""
    if isinstance(mission, tuple):
        goal_str, constr_str = mission
    else:
        parts = mission.split(" and avoid ", 1)
        goal_str, constr_str = (parts[0], "avoid " + parts[1]) if len(parts) == 2 else (mission, None)

    with torch.no_grad():
        g_emb = mission_encoder(goal_str).to(device)
        c_emb = mission_encoder(constr_str).to(device) if constr_str else torch.zeros_like(g_emb)

        if hypernetwork is not None:
            theta_flat    = parameters_to_vector(list(policy_h.parameters()))
            combined_inp  = torch.cat([theta_flat.unsqueeze(0), g_emb, c_emb], dim=-1)
            theta_tensors = hypernetwork(combined_inp)
            names = list(dict(policy_h.named_parameters()).keys())
            return OrderedDict((n, t.squeeze(0)) for n, t in zip(names, theta_tensors))
        else:
            return OrderedDict(policy_h.named_parameters())


# ── Single-episode rollout ─────────────────────────────────────────────────────
def rollout(policy, params, preproc=sampler_lang.preprocess_obs, seed=None):
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', max_steps)
    while not done and steps < env_max:
        obs_t = torch.from_numpy(preproc(obs)[None]).float().to(device)
        with torch.no_grad():
            action = policy(obs_t, params=params).sample().item()
        obs, reward, terminated, truncated, info = env.step(action)
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True
    return steps, success, viols


def rollout_random(seed=None):
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', max_steps)
    while not done and steps < env_max:
        obs, reward, terminated, truncated, info = env.step(env.action_space.sample())
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True
    return steps, success, viols


# ── Select test tasks ──────────────────────────────────────────────────────────
test_tasks = random.sample(all_missions, min(N, len(all_missions)))

# ── Run evaluation ─────────────────────────────────────────────────────────────
METHODS = []
if clamaml_ready:  METHODS.append(("C-LA-MAML",   policy_c, _params_clamaml))
if unified_ready:  METHODS.append(("Unified",      policy_u, _params_unified))
if hyper_ready:    METHODS.append(("Hyper",        policy_h, _params_hyper))
METHODS.append(("Random", None, None))

# per-method stats: list of (sr, avg_steps, avg_viols) per task
per_task = {m[0]: [] for m in METHODS}

COL_W = 14
header = f"{'Mission':<45}" + "".join(f"{'  SR%/Steps/Viols':>{COL_W*3}}" for m,_,__ in METHODS)
print(f"\n{'Mission':<45} " +
      " | ".join(f"{'SR%':>5} {'Steps':>6} {'Viols':>5}" for m,_,__ in METHODS))
print(f"  ({'  |  '.join(m[0] for m in METHODS)})")
print("-" * (45 + 22 * len(METHODS)))

for mission in test_tasks:
    env.reset_task(mission)
    mstr = f"{mission[0]} and {mission[1]}" if isinstance(mission, tuple) else mission
    row  = f"{mstr[:44]:<45}"

    ep_seeds = [random.randint(0, 1000000) for _ in range(EP)]

    for (mname, policy, get_params) in METHODS:
        ep_steps, ep_succ, ep_viols = [], [], []
        if mname == "Random":
            for ep in range(EP):
                env.reset_task(mission)
                s, ok, v = rollout_random(seed=ep_seeds[ep])
                ep_steps.append(s); ep_succ.append(ok); ep_viols.append(v)
                print(f"  [{mname}] Ep {ep+1}/{EP} - Success: {ok}, Steps: {s}, Viols: {v}")
        else:
            params = get_params(mission)
            for ep in range(EP):
                env.reset_task(mission)
                s, ok, v = rollout(policy, params, seed=ep_seeds[ep])
                ep_steps.append(s); ep_succ.append(ok); ep_viols.append(v)
                print(f"  [{mname}] Ep {ep+1}/{EP} - Success: {ok}, Steps: {s}, Viols: {v}")

        sr    = np.mean(ep_succ)
        avgst = np.mean(ep_steps)
        avgv  = np.mean(ep_viols)
        per_task[mname].append((sr, avgst, avgv))
        row += f" | {sr*100:>4.0f}% {avgst:>6.1f} {avgv:>5.1f}"

    print(row)

# ── Summary ────────────────────────────────────────────────────────────────────
print("\n" + "="*65)
print(f"{'Method':<18} | {'Avg SR%':>7} | {'Avg Steps':>9} | {'Avg Viols':>9}")
print("-"*65)
for (mname, _, __) in METHODS:
    data = per_task[mname]
    srs   = [x[0] for x in data]
    steps = [x[1] for x in data]
    viols = [x[2] for x in data]
    print(f"{mname:<18} | {np.mean(srs)*100:>6.1f}% | {np.mean(steps):>9.1f} | {np.mean(viols):>9.2f}")
print("="*65)

# ── Excel logging ──────────────────────────────────────────────────────────────
xlsx_path = "evaluation_results_all.xlsx"
if os.path.exists(xlsx_path):
    wb = load_workbook(xlsx_path)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

# Summary sheet (append one row per run)
summary_name = "Summary"
if summary_name not in wb.sheetnames:
    ws_sum = wb.create_sheet(summary_name)
    ws_sum.append(["Env", "Room", "Dists", "Steps", "Delta",
                   "C-SR%","C-Steps","C-Viols",
                   "U-SR%","U-Steps","U-Viols",
                   "H-SR%","H-Steps","H-Viols",
                   "Rand-SR%","Rand-Steps","Rand-Viols"])
else:
    ws_sum = wb[summary_name]

def _agg(mname):
    data = per_task.get(mname, [])
    if not data:
        return "", "", ""
    return (round(np.mean([x[0] for x in data])*100, 1),
            round(np.mean([x[1] for x in data]), 1),
            round(np.mean([x[2] for x in data]), 2))

ws_sum.append([env_name, room_size, num_dists, max_steps, delta_theta,
               *_agg("C-LA-MAML"), *_agg("Unified"), *_agg("Hyper"), *_agg("Random")])

# Per-mission sheet (overwritten per run)
sheet_name = (env_name + "_Missions")[:31]
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws_m = wb.create_sheet(sheet_name)
header_row = ["Mission"]
for (mname, _, __) in METHODS:
    header_row += [f"{mname} SR%", f"{mname} Steps", f"{mname} Viols"]
ws_m.append(header_row)

for i, mission in enumerate(test_tasks):
    mstr = f"{mission[0]} and {mission[1]}" if isinstance(mission, tuple) else mission
    row = [mstr]
    for (mname, _, __) in METHODS:
        sr, st, vl = per_task[mname][i]
        row += [round(sr*100,1), round(st,1), round(vl,2)]
    ws_m.append(row)

# OVERALL row
overall_row = ["OVERALL"]
for (mname, _, __) in METHODS:
    sr, st, vl = _agg(mname)
    overall_row += [sr, st, vl]
ws_m.append(overall_row)

# Make the last row bold
bold_font = Font(bold=True)
for cell in ws_m[ws_m.max_row]:
    cell.font = bold_font

wb.save(xlsx_path)
print(f"\nResults saved → {xlsx_path}  (sheets: '{summary_name}', '{sheet_name}')")
